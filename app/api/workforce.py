"""
API роутер модуля управления численностью.
v2: объекты + план от стройки
"""
import io
from datetime import date, datetime
from typing import List, Optional
from uuid import UUID

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ..database import get_session
from ..models.workforce import (
    Project, ProjectObject, WorkforceNorm, BudgetPeriod, BudgetItem,
    HeadcountFact, HeadcountPlan,
    WfContractor, WfContractorAssignment,
    Challenge, ChallengeItem, MobilizationPlan, MobilizationCheckpoint,
    Violation, ArticleMapping,
)
from ..schemas.workforce import (
    ProjectCreate, ProjectOut,
    ProjectObjectCreate, ProjectObjectOut,
    NormCreate, NormOut,
    BudgetPeriodOut,
    HeadcountFactCreate, HeadcountFactOut,
    HeadcountPlanCreate, HeadcountPlanOut,
    DashboardResponse, ProjectDetailResponse,
    ContractorCreate, ContractorOut,
    ForecastResponse,
    ContractorHeadcountRow,
    ChallengeCreate, ChallengeOut, ChallengeUpdate,
    ChallengeItemOut, MobilizationPlanBulkCreate, MobilizationPlanOut,
    ViolationCreate, ViolationOut, ViolationUpdate, ViolationScanResult,
    SystemProblemsResponse,
    ContractorRatingRow,
    ArticleMappingCreate, ArticleMappingOut, ArticleMappingBulk, UnmappedArticleOut,
)
from ..services.workforce import (
    calc_dashboard, calc_project_detail,
    calc_forecast, calc_object_contractors,
    calc_system_problems, calc_contractor_rating,
    check_challenge_checkpoints, scan_violations, enrich_violations, auto_escalate_violations,
)

router = APIRouter(prefix="/api/v1/workforce", tags=["workforce"])


# ── Projects ──────────────────────────────────────────────────────────────────

@router.get("/projects", response_model=List[ProjectOut])
async def list_projects(session: AsyncSession = Depends(get_session)):
    result = await session.execute(select(Project).order_by(Project.name))
    return result.scalars().all()


@router.post("/projects", response_model=ProjectOut, status_code=201)
async def create_project(data: ProjectCreate, session: AsyncSession = Depends(get_session)):
    proj = Project(**data.model_dump())
    session.add(proj)
    await session.commit()
    await session.refresh(proj)
    return proj


# ── Project Objects ───────────────────────────────────────────────────────────

@router.get("/projects/{project_id}/objects", response_model=List[ProjectObjectOut])
async def list_objects(project_id: UUID, session: AsyncSession = Depends(get_session)):
    result = await session.execute(
        select(ProjectObject).where(ProjectObject.project_id == project_id).order_by(ProjectObject.name)
    )
    return result.scalars().all()


@router.post("/projects/{project_id}/objects", response_model=ProjectObjectOut, status_code=201)
async def create_object(project_id: UUID, data: ProjectObjectCreate, session: AsyncSession = Depends(get_session)):
    proj = (await session.execute(select(Project).where(Project.id == project_id))).scalar_one_or_none()
    if not proj:
        raise HTTPException(404, "Проект не найден")
    obj = ProjectObject(project_id=project_id, **data.model_dump())
    session.add(obj)
    await session.commit()
    await session.refresh(obj)
    return obj


# ── Norms ─────────────────────────────────────────────────────────────────────

@router.get("/norms", response_model=List[NormOut])
async def list_norms(
    project_class: Optional[str] = Query(None),
    session: AsyncSession = Depends(get_session),
):
    q = select(WorkforceNorm).order_by(WorkforceNorm.work_type, WorkforceNorm.project_class)
    if project_class:
        q = q.where(WorkforceNorm.project_class == project_class)
    return (await session.execute(q)).scalars().all()


@router.post("/norms", response_model=NormOut, status_code=201)
async def upsert_norm(data: NormCreate, session: AsyncSession = Depends(get_session)):
    result = await session.execute(
        select(WorkforceNorm)
        .where(WorkforceNorm.work_type == data.work_type, WorkforceNorm.project_class == data.project_class)
    )
    norm = result.scalar_one_or_none()
    if norm:
        for field, value in data.model_dump().items():
            setattr(norm, field, value)
    else:
        norm = WorkforceNorm(**data.model_dump())
        session.add(norm)
    await session.commit()
    await session.refresh(norm)
    return norm


# ── Budget upload ─────────────────────────────────────────────────────────────

import re
ARTICLE_CODE_RE = re.compile(r"^(\d{2}-\d{2}(?:-\d{2}){0,2}-Р)\s*,\s*(.+?)\*?$")
# Месяцы РФ → номер
RU_MONTHS = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4, "май": 5, "июнь": 6,
    "июль": 7, "август": 8, "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}


def _parse_month_header(text: str) -> Optional[date]:
    """Парсит «Январь 2026» → date(2026, 1, 1). None если не подходит."""
    if not text:
        return None
    parts = str(text).strip().lower().split()
    if len(parts) != 2:
        return None
    m_name, year_str = parts
    m = RU_MONTHS.get(m_name)
    if not m:
        return None
    try:
        year = int(year_str)
    except ValueError:
        return None
    return date(year, m, 1)


def _parse_article(text: str) -> Optional[tuple[str, str]]:
    """Из «40-02-01-Р, Устройство свайного поля и шпунтов*» → ('40-02-01-Р', 'Устройство свайного поля и шпунтов')."""
    if not text:
        return None
    s = str(text).strip()
    m = ARTICLE_CODE_RE.match(s)
    if not m:
        return None
    code = m.group(1)
    label = m.group(2).strip().rstrip("*").strip()
    return code, label


@router.post("/upload-budget", response_model=dict)
async def upload_budget(file: UploadFile = File(...), session: AsyncSession = Depends(get_session)):
    """
    Загрузка иерархического бюджета (выгрузка из 1С: «Утверждённый бюджет»).
    Структура (по outline_level в Excel):
      L0/L1: «БДР» (итог)
      L1   : Проект
      L2   : Объект (Корпус)
      L3   : «БДР, БДР»
      L4-L7: иерархия статей (40-Р → 40-02-Р → 40-02-01-Р → 40-02-01-01-Р)
      L7   : также контрагенты и договоры (отфильтровываются)

    Берётся самый детальный уровень статьи как `detailed_article`. Для каждой
    статьи создаются BudgetItem на каждый месяц колонок-периодов.

    Соответствие detailed_article → work_type подтягивается из ArticleMapping.
    Если маппинга нет — в work_type записывается сам код, статья помечается
    как «не смаппленная» (видно в /article-mapping/unmapped).
    """
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Не удалось прочитать Excel")

    ws = wb.active

    # Найти строку заголовков с месяцами и колонку «УТВЕРЖДЕННЫЙ БЮДЖЕТ»
    header_row_idx = None
    month_cols: dict[int, date] = {}
    total_col: Optional[int] = None
    for r in range(1, min(20, ws.max_row + 1)):
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        for i, v in enumerate(row, 1):
            if v is None:
                continue
            text = str(v).strip()
            pm = _parse_month_header(text)
            if pm:
                month_cols[i] = pm
                header_row_idx = r
            elif text.upper() == "УТВЕРЖДЕННЫЙ БЮДЖЕТ":
                total_col = i
                header_row_idx = r
        if month_cols:
            break

    if not month_cols:
        raise HTTPException(400, "Не найдены колонки с месяцами в шапке файла")

    data_start = (header_row_idx or 6) + 2  # пропускаем подзаголовок «Сумма (сценарий)»

    # Загружаем существующий маппинг статей
    existing_mappings = {
        am.article_code: am.work_type
        for am in (await session.execute(select(ArticleMapping))).scalars().all()
    }

    # Иерархический проход
    current_project: Optional[str] = None
    current_object: Optional[str] = None
    current_article_code: Optional[str] = None
    current_article_label: Optional[str] = None
    current_article_level: int = -1  # outline_level статьи, чтобы знать когда сбросить

    project_cache: dict[str, Project] = {}
    object_cache: dict[tuple[str, str], ProjectObject] = {}
    items_to_add: list[dict] = []
    unmapped_articles: dict[str, str] = {}

    for r in range(data_start, ws.max_row + 1):
        row_dim = ws.row_dimensions.get(r)
        level = row_dim.outline_level if row_dim else 0
        row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        label_raw = row[0]
        if label_raw is None:
            continue
        label = str(label_raw).strip()
        if not label:
            continue

        # L1 — проект (между уровнями БДР и Объект)
        if level == 1 and label.upper() != "БДР":
            current_project = label
            current_object = None
            current_article_code = None
            current_article_label = None
            current_article_level = -1
            continue

        # L2 — объект строительства
        if level == 2:
            current_object = label
            current_article_code = None
            current_article_label = None
            current_article_level = -1
            continue

        # L3 — «БДР, БДР» — пропускаем
        if level == 3:
            continue

        # Любая глубже L3 — может быть статья или строка под статью (контрагент/договор)
        if level >= 4:
            parsed = _parse_article(label)
            if parsed is not None:
                # Это строка-статья. Берём её как самую детальную на текущем уровне.
                code, art_label = parsed
                # Если новая статья «глубже» предыдущей — она самая детальная
                if level >= current_article_level:
                    current_article_code = code
                    current_article_label = art_label
                    current_article_level = level
                else:
                    # Поднялись на уровень выше — сбросили
                    current_article_code = code
                    current_article_label = art_label
                    current_article_level = level

                # Записываем суммы по месяцам для этой статьи
                if current_project and current_object:
                    month_values = {}
                    for col_idx, pm in month_cols.items():
                        v = row[col_idx - 1] if col_idx - 1 < len(row) else None
                        if v is None or v == 0:
                            continue
                        try:
                            amount = abs(float(v))  # бюджет идёт со знаком минус — приводим к плюсу
                        except (TypeError, ValueError):
                            continue
                        if amount > 0:
                            month_values[pm] = amount
                    if month_values:
                        work_type = existing_mappings.get(code, code)
                        if code not in existing_mappings:
                            unmapped_articles[code] = art_label
                        items_to_add.append({
                            "project": current_project,
                            "object": current_object,
                            "article_code": code,
                            "article_label": art_label,
                            "work_type": work_type,
                            "months": month_values,
                        })
            else:
                # Это контрагент или договор — пропускаем (суммы у них дублируют статью выше)
                continue

    if not items_to_add:
        raise HTTPException(400, "Не найдено ни одной строки бюджета с данными")

    # Удалить старые BudgetPeriod для затронутых проектов
    affected_projects = {it["project"] for it in items_to_add}
    for pname in affected_projects:
        proj = (await session.execute(select(Project).where(Project.name == pname))).scalar_one_or_none()
        if proj:
            old_periods = (await session.execute(
                select(BudgetPeriod).where(BudgetPeriod.project_id == proj.id)
            )).scalars().all()
            for op in old_periods:
                await session.delete(op)
    await session.flush()

    # Кэш проектов и объектов
    for pname in affected_projects:
        proj = (await session.execute(select(Project).where(Project.name == pname))).scalar_one_or_none()
        if not proj:
            proj = Project(name=pname)
            session.add(proj)
            await session.flush()
        project_cache[pname] = proj

    # Создаём BudgetPeriod на каждый (project, month)
    period_cache: dict[tuple[str, date], BudgetPeriod] = {}
    upload_ts = datetime.utcnow()
    for it in items_to_add:
        proj = project_cache[it["project"]]
        for pm in it["months"]:
            key = (it["project"], pm)
            if key not in period_cache:
                bp = BudgetPeriod(project_id=proj.id, period_month=pm, upload_date=upload_ts)
                session.add(bp)
                await session.flush()
                period_cache[key] = bp
    await session.flush()

    # Объекты
    for it in items_to_add:
        proj = project_cache[it["project"]]
        obj_key = (it["project"], it["object"])
        if obj_key not in object_cache:
            obj = (await session.execute(
                select(ProjectObject).where(
                    ProjectObject.project_id == proj.id,
                    ProjectObject.name == it["object"]
                )
            )).scalar_one_or_none()
            if not obj:
                obj = ProjectObject(project_id=proj.id, name=it["object"])
                session.add(obj)
                await session.flush()
            object_cache[obj_key] = obj
    await session.flush()

    # Создаём BudgetItem
    items_count = 0
    for it in items_to_add:
        obj = object_cache[(it["project"], it["object"])]
        for pm, amount in it["months"].items():
            bp = period_cache[(it["project"], pm)]
            session.add(BudgetItem(
                budget_period_id=bp.id,
                object_id=obj.id,
                work_type=it["work_type"],
                detailed_article=f"{it['article_code']}, {it['article_label']}",
                bdr_amount=amount,
                management_completion_amount=0,
            ))
            items_count += 1

    await session.commit()

    return {
        "projects_imported": len(affected_projects),
        "objects_imported": len(object_cache),
        "periods_imported": len(period_cache),
        "items_imported": items_count,
        "unmapped_articles_count": len(unmapped_articles),
        "unmapped_articles": [
            {"article_code": code, "article_label": label}
            for code, label in list(unmapped_articles.items())[:50]
        ],
    }


# ── Article Mapping CRUD ──────────────────────────────────────────────────────

@router.get("/article-mapping", response_model=List[ArticleMappingOut])
async def list_article_mappings(session: AsyncSession = Depends(get_session)):
    result = await session.execute(select(ArticleMapping).order_by(ArticleMapping.article_code))
    return result.scalars().all()


@router.post("/article-mapping", response_model=ArticleMappingOut, status_code=201)
async def upsert_article_mapping(data: ArticleMappingCreate, session: AsyncSession = Depends(get_session)):
    am = (await session.execute(
        select(ArticleMapping).where(ArticleMapping.article_code == data.article_code)
    )).scalar_one_or_none()
    if am:
        am.article_label = data.article_label
        am.work_type = data.work_type
    else:
        am = ArticleMapping(**data.model_dump())
        session.add(am)
    await session.commit()
    await session.refresh(am)
    return am


@router.post("/article-mapping/bulk", response_model=dict)
async def bulk_article_mapping(data: ArticleMappingBulk, session: AsyncSession = Depends(get_session)):
    upserted = 0
    for item in data.items:
        am = (await session.execute(
            select(ArticleMapping).where(ArticleMapping.article_code == item.article_code)
        )).scalar_one_or_none()
        if am:
            am.article_label = item.article_label
            am.work_type = item.work_type
        else:
            session.add(ArticleMapping(**item.model_dump()))
        upserted += 1
    await session.commit()
    return {"upserted": upserted}


@router.delete("/article-mapping/{article_code}", status_code=204)
async def delete_article_mapping(article_code: str, session: AsyncSession = Depends(get_session)):
    am = (await session.execute(
        select(ArticleMapping).where(ArticleMapping.article_code == article_code)
    )).scalar_one_or_none()
    if not am:
        raise HTTPException(404, "Маппинг не найден")
    await session.delete(am)
    await session.commit()


@router.get("/article-mapping/unmapped", response_model=List[UnmappedArticleOut])
async def list_unmapped_articles(session: AsyncSession = Depends(get_session)):
    """Статьи из загруженного бюджета, для которых нет маппинга."""
    mapped_codes = {am.article_code for am in (await session.execute(select(ArticleMapping))).scalars().all()}
    items = (await session.execute(select(BudgetItem).where(BudgetItem.detailed_article.is_not(None)))).scalars().all()
    counter: dict[str, dict] = {}
    for it in items:
        if not it.detailed_article:
            continue
        code = it.detailed_article.split(",", 1)[0].strip()
        if code in mapped_codes:
            continue
        if code not in counter:
            label = it.detailed_article.split(",", 1)[1].strip() if "," in it.detailed_article else ""
            counter[code] = {"article_code": code, "article_label": label, "occurrences": 0}
        counter[code]["occurrences"] += 1
    return sorted(counter.values(), key=lambda x: -x["occurrences"])


# ── Dashboard ─────────────────────────────────────────────────────────────────

@router.get("/dashboard", response_model=DashboardResponse)
async def get_dashboard(session: AsyncSession = Depends(get_session)):
    return await calc_dashboard(session)


@router.get("/project/{project_id}", response_model=ProjectDetailResponse)
async def get_project_detail_endpoint(
    project_id: UUID,
    period_month: Optional[date] = Query(None),
    session: AsyncSession = Depends(get_session),
):
    proj = (await session.execute(select(Project).where(Project.id == project_id))).scalar_one_or_none()
    if not proj:
        raise HTTPException(404, "Проект не найден")
    return await calc_project_detail(session, proj, period_month)


# ── Headcount ─────────────────────────────────────────────────────────────────

@router.post("/headcount", response_model=HeadcountFactOut, status_code=201)
async def submit_headcount(data: HeadcountFactCreate, session: AsyncSession = Depends(get_session)):
    proj = (await session.execute(select(Project).where(Project.id == data.project_id))).scalar_one_or_none()
    if not proj:
        raise HTTPException(404, "Проект не найден")
    fact = HeadcountFact(**data.model_dump())
    session.add(fact)
    await session.commit()
    await session.refresh(fact)
    return fact


@router.post("/headcount-plan", response_model=HeadcountPlanOut, status_code=201)
async def submit_headcount_plan(data: HeadcountPlanCreate, session: AsyncSession = Depends(get_session)):
    proj = (await session.execute(select(Project).where(Project.id == data.project_id))).scalar_one_or_none()
    if not proj:
        raise HTTPException(404, "Проект не найден")
    plan = HeadcountPlan(**data.model_dump())
    session.add(plan)
    await session.commit()
    await session.refresh(plan)
    return plan


# ── WfContractors ───────────────────────────────────────────────────────────────

@router.get("/contractors", response_model=List[ContractorOut])
async def list_contractors(session: AsyncSession = Depends(get_session)):
    result = await session.execute(select(WfContractor).order_by(WfContractor.name))
    return result.scalars().all()


@router.post("/contractors", response_model=ContractorOut, status_code=201)
async def create_contractor(data: ContractorCreate, session: AsyncSession = Depends(get_session)):
    c = WfContractor(**data.model_dump())
    session.add(c)
    await session.commit()
    await session.refresh(c)
    return c


@router.get("/contractors/{contractor_id}", response_model=ContractorOut)
async def get_contractor(contractor_id: UUID, session: AsyncSession = Depends(get_session)):
    c = (await session.execute(select(WfContractor).where(WfContractor.id == contractor_id))).scalar_one_or_none()
    if not c:
        raise HTTPException(404, "Подрядчик не найден")
    return c


# ── Forecast ──────────────────────────────────────────────────────────────────

@router.get("/project/{project_id}/forecast", response_model=ForecastResponse)
async def get_forecast(project_id: UUID, session: AsyncSession = Depends(get_session)):
    proj = (await session.execute(select(Project).where(Project.id == project_id))).scalar_one_or_none()
    if not proj:
        raise HTTPException(404, "Проект не найден")
    return await calc_forecast(session, project_id)


# ── Object contractors ────────────────────────────────────────────────────────

@router.get("/objects/{object_id}/contractors", response_model=List[ContractorHeadcountRow])
async def get_object_contractors(object_id: UUID, session: AsyncSession = Depends(get_session)):
    obj = (await session.execute(select(ProjectObject).where(ProjectObject.id == object_id))).scalar_one_or_none()
    if not obj:
        raise HTTPException(404, "Объект не найден")
    return await calc_object_contractors(session, object_id)


# ── Challenges ────────────────────────────────────────────────────────────────

@router.post("/challenges", response_model=ChallengeOut, status_code=201)
async def create_challenge(data: ChallengeCreate, session: AsyncSession = Depends(get_session)):
    challenge = Challenge(
        project_id=data.project_id,
        object_id=data.object_id,
        period_month=data.period_month,
        comment=data.comment,
    )
    session.add(challenge)
    await session.flush()
    for item_data in data.items:
        session.add(ChallengeItem(challenge_id=challenge.id, **item_data.model_dump()))
    await session.commit()
    await session.refresh(challenge)
    # Load items
    items = (await session.execute(
        select(ChallengeItem).where(ChallengeItem.challenge_id == challenge.id)
    )).scalars().all()
    return ChallengeOut(
        id=challenge.id,
        project_id=challenge.project_id,
        object_id=challenge.object_id,
        period_month=challenge.period_month,
        status=challenge.status,
        created_at=challenge.created_at,
        approved_by=challenge.approved_by,
        approved_at=challenge.approved_at,
        comment=challenge.comment,
        items=[ChallengeItemOut.model_validate(i) for i in items],
    )


@router.get("/challenges", response_model=List[ChallengeOut])
async def list_challenges(
    project_id: Optional[UUID] = Query(None),
    period: Optional[date] = Query(None),
    session: AsyncSession = Depends(get_session),
):
    q = select(Challenge).order_by(Challenge.created_at.desc())
    if project_id:
        q = q.where(Challenge.project_id == project_id)
    if period:
        q = q.where(Challenge.period_month == period)
    challenges = (await session.execute(q)).scalars().all()
    result = []
    for ch in challenges:
        items = (await session.execute(
            select(ChallengeItem).where(ChallengeItem.challenge_id == ch.id)
        )).scalars().all()
        result.append(ChallengeOut(
            id=ch.id, project_id=ch.project_id, object_id=ch.object_id,
            period_month=ch.period_month, status=ch.status, created_at=ch.created_at,
            approved_by=ch.approved_by, approved_at=ch.approved_at, comment=ch.comment,
            items=[ChallengeItemOut.model_validate(i) for i in items],
        ))
    return result


@router.get("/challenges/{challenge_id}", response_model=ChallengeOut)
async def get_challenge(challenge_id: UUID, session: AsyncSession = Depends(get_session)):
    ch = (await session.execute(select(Challenge).where(Challenge.id == challenge_id))).scalar_one_or_none()
    if not ch:
        raise HTTPException(404, "Челлендж не найден")
    items = (await session.execute(
        select(ChallengeItem).where(ChallengeItem.challenge_id == ch.id)
    )).scalars().all()
    items_out = []
    for item in items:
        plans = (await session.execute(
            select(MobilizationPlan).where(MobilizationPlan.challenge_item_id == item.id)
        )).scalars().all()
        items_out.append(ChallengeItemOut(
            **{k: getattr(item, k) for k in ["id", "challenge_id", "work_type", "system_baseline",
                                              "requested_count", "approved_count", "requires_mobilization_plan"]},
            mobilization_plans=[MobilizationPlanOut.model_validate(p) for p in plans],
        ))
    return ChallengeOut(
        id=ch.id, project_id=ch.project_id, object_id=ch.object_id,
        period_month=ch.period_month, status=ch.status, created_at=ch.created_at,
        approved_by=ch.approved_by, approved_at=ch.approved_at, comment=ch.comment,
        items=items_out,
    )


@router.patch("/challenges/{challenge_id}", response_model=ChallengeOut)
async def update_challenge(
    challenge_id: UUID, data: ChallengeUpdate,
    session: AsyncSession = Depends(get_session),
):
    ch = (await session.execute(select(Challenge).where(Challenge.id == challenge_id))).scalar_one_or_none()
    if not ch:
        raise HTTPException(404, "Челлендж не найден")
    if data.status is not None:
        ch.status = data.status
        if data.status in ("approved", "rejected"):
            ch.approved_at = datetime.utcnow()
    if data.approved_by is not None:
        ch.approved_by = data.approved_by
    if data.comment is not None:
        ch.comment = data.comment
    await session.commit()
    return await get_challenge(challenge_id, session)


@router.post("/challenges/{challenge_id}/mobilization", response_model=List[MobilizationPlanOut], status_code=201)
async def add_mobilization_plan(
    challenge_id: UUID, data: MobilizationPlanBulkCreate,
    session: AsyncSession = Depends(get_session),
):
    ch = (await session.execute(select(Challenge).where(Challenge.id == challenge_id))).scalar_one_or_none()
    if not ch:
        raise HTTPException(404, "Челлендж не найден")
    ci = (await session.execute(
        select(ChallengeItem).where(
            ChallengeItem.id == data.challenge_item_id,
            ChallengeItem.challenge_id == challenge_id,
        )
    )).scalar_one_or_none()
    if not ci:
        raise HTTPException(404, "Строка челленджа не найдена")
    created = []
    for p in data.plans:
        mp = MobilizationPlan(challenge_item_id=ci.id, **p.model_dump())
        session.add(mp)
        await session.flush()
        created.append(mp)
    await session.commit()
    return [MobilizationPlanOut.model_validate(mp) for mp in created]


@router.post("/challenges/{challenge_id}/checkpoints/check")
async def check_checkpoints(challenge_id: UUID, session: AsyncSession = Depends(get_session)):
    updated = await check_challenge_checkpoints(session, challenge_id)
    return {"updated": updated}


# ── Violations ────────────────────────────────────────────────────────────────

@router.get("/violations", response_model=List[ViolationOut])
async def list_violations(
    project_id: Optional[UUID] = Query(None),
    object_id: Optional[UUID] = Query(None),
    resolved: Optional[bool] = Query(None),
    violation_type: Optional[str] = Query(None),
    session: AsyncSession = Depends(get_session),
):
    q = select(Violation).order_by(Violation.violation_date.desc())
    if project_id:
        q = q.where(Violation.project_id == project_id)
    if object_id:
        q = q.where(Violation.object_id == object_id)
    if resolved is not None:
        q = q.where(Violation.resolved == resolved)
    if violation_type:
        q = q.where(Violation.violation_type == violation_type)
    violations = (await session.execute(q)).scalars().all()
    return await enrich_violations(session, violations)




@router.post("/violations/scan", response_model=ViolationScanResult)
async def run_violations_scan(session: AsyncSession = Depends(get_session)):
    """Автоматическая проверка и создание нарушений по текущим данным"""
    # 1. Auto-escalate existing violations by timer
    escalated = await auto_escalate_violations(session)
    # 2. Scan for new violations
    new_violations = await scan_violations(session)
    enriched = await enrich_violations(session, new_violations)
    return ViolationScanResult(created=len(enriched), violations=enriched)

@router.post("/violations", response_model=ViolationOut, status_code=201)
async def create_violation(data: ViolationCreate, session: AsyncSession = Depends(get_session)):
    v = Violation(**data.model_dump())
    session.add(v)
    await session.commit()
    await session.refresh(v)
    return v


@router.patch("/violations/{violation_id}", response_model=ViolationOut)
async def update_violation(
    violation_id: UUID, data: ViolationUpdate,
    session: AsyncSession = Depends(get_session),
):
    v = (await session.execute(select(Violation).where(Violation.id == violation_id))).scalar_one_or_none()
    if not v:
        raise HTTPException(404, "Нарушение не найдено")
    if data.escalated is not None:
        v.escalated = data.escalated
    if data.escalated_to is not None:
        v.escalated_to = data.escalated_to
    if data.resolved is not None:
        v.resolved = data.resolved
    await session.commit()
    await session.refresh(v)
    return v


# ── Analytics ─────────────────────────────────────────────────────────────────

@router.get("/analytics/system-problems", response_model=SystemProblemsResponse)
async def get_system_problems(
    threshold_pct: float = Query(50.0),
    min_objects: int = Query(3),
    session: AsyncSession = Depends(get_session),
):
    return await calc_system_problems(session, threshold_pct, min_objects)


@router.get("/analytics/contractor-rating", response_model=List[ContractorRatingRow])
async def get_contractor_rating(session: AsyncSession = Depends(get_session)):
    return await calc_contractor_rating(session)
